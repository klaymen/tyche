"""Regenerate sample_input.xlsx with Project Group and Probability columns."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

months = [
    "2026-04","2026-05","2026-06","2026-07","2026-08","2026-09",
    "2026-10","2026-11","2026-12","2027-01","2027-02","2027-03",
]

# [Employee, Project, Project Group, Probability, Hourly Rate, ...monthly FTE...]
rows = [
    ["Alice Martin",  "Project Alpha", "Core Products",  1.00, 50, 0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6],
    ["Alice Martin",  "Project Beta",  "Client Work",    1.00, 50, 0.4,0.4,0.4,0.4,0.4,0.4,0.4,0.4,0.4,0.4,0.4,0.4],
    ["Bob Chen",      "Project Beta",  "Client Work",    1.00, 50, 0.5,0.5,0.5,0.5,0.5,0.5,0.3,0.3,0.3,0.3,0.3,0.3],
    ["Bob Chen",      "Project Gamma", "Internal / Ops", 0.55, 50, 0.5,0.5,0.5,0.5,0.7,0.7,0.7,0.7,0.7,0.7,0.7,0.7],
    ["Carol Singh",   "Project Alpha", "Core Products",  1.00, 50, 1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0],
    ["David Lee",     "Project Beta",  "Client Work",    1.00, 50, 0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5],
    ["Eva Torres",    "Project Gamma", "Internal / Ops", 1.00, 50, 0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3],
    ["Frank Müller",  "Project Alpha", "Core Products",  1.00, 50, 0.6,0.6,0.5,0.5,0.4,0.4,0.3,0.3,0.2,0.2,0.1,0.1],
    ["Frank Müller",  "Project Beta",  "Client Work",    1.00, 50, 0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3],
    ["Frank Müller",  "Project Gamma", "Internal / Ops", 0.50, 50, 0.1,0.1,0.2,0.2,0.3,0.3,0.4,0.4,0.5,0.5,0.6,0.6],
    ["Grace Kim",     "Project Gamma", "Internal / Ops", 1.00, 50, 1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0],
    ["Henry Osei",    "Project Beta",  "Client Work",    1.00, 50, 1.0,1.0,1.0,0.5,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0],
    ["Henry Osei",    "Project Alpha", "Core Products",  1.00, 50, 0.0,0.0,0.0,0.5,1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.0],
    ["Iris Dubois",   "Project Alpha", "Core Products",  1.00, 50, 0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5,0.5],
    ["Iris Dubois",   "Project Beta",  "Client Work",    0.70, 50, 0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3,0.3],
    ["James Patel",   "Project Gamma", "Internal / Ops", 1.00, 50, 0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6],
    ["Lena Brandt",   "Project Beta",  "Client Work",    1.00, 50, 0.0,0.0,0.0,0.0,0.0,0.0,0.4,0.4,0.4,0.4,0.4,0.4],
    ["Marco Silva",   "Project Alpha", "Core Products",  1.00, 50, 0.7,0.7,0.6,0.6,0.5,0.5,0.5,0.5,0.6,0.6,0.7,0.7],
    ["Marco Silva",   "Project Gamma", "Internal / Ops", 1.00, 50, 0.3,0.3,0.4,0.4,0.5,0.5,0.5,0.5,0.4,0.4,0.3,0.3],
    # Project Delta — Client Work (same group as Beta, adds a second project to the group)
    ["Nina Walsh",    "Project Delta", "Client Work",    0.75, 50, 0.5,0.5,0.6,0.6,0.7,0.7,0.8,0.8,0.8,0.8,0.7,0.7],
    ["Oscar Grant",   "Project Delta", "Client Work",    0.80, 50, 0.8,0.8,0.8,0.8,0.6,0.6,0.4,0.4,0.2,0.2,0.0,0.0],
    ["Lena Brandt",   "Project Delta", "Client Work",    1.00, 50, 0.6,0.6,0.6,0.6,0.6,0.6,0.0,0.0,0.0,0.0,0.0,0.0],
    ["Alice Martin",  "Project Delta", "Client Work",    1.00, 50, 0.0,0.0,0.0,0.2,0.2,0.2,0.2,0.2,0.2,0.0,0.0,0.0],
    ["David Lee",     "Project Delta", "Client Work",    1.00, 50, 0.0,0.0,0.0,0.0,0.3,0.3,0.3,0.3,0.3,0.3,0.5,0.5],
]

wb = Workbook()
ws = wb.active

# Row 1: "Months" label starting at the first month column (col 6)
ws.cell(1, 6).value = "Months"
ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=5 + len(months))
ws.cell(1, 6).font = Font(bold=True)
ws.cell(1, 6).alignment = Alignment(horizontal="center")

# Row 2: column headers
headers = ["Employee", "Project", "Project Group", "Probability", "Hourly Rate"] + months
for col, h in enumerate(headers, 1):
    ws.cell(2, col).value = h
    ws.cell(2, col).font = Font(bold=True)

# Data rows starting at row 3
for ri, row_data in enumerate(rows, 3):
    for ci, val in enumerate(row_data, 1):
        ws.cell(ri, ci).value = val

wb.save("sample_input.xlsx")
print(f"Saved {len(rows)} rows.")

# Verify
import pandas as pd
df = pd.read_excel("sample_input.xlsx", header=1)
print(df[["Employee", "Project", "Project Group", "Probability"]].to_string())
